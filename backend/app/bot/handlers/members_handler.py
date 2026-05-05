import io
import os
from datetime import datetime, timedelta, timezone
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile
from aiogram.filters import Command
from sqlalchemy import select, func

from backend.app.db.session import async_session
from backend.app.db.models import User, Membership, Channel

router = Router()

ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x]
IST = timezone(timedelta(hours=5, minutes=30))
PAGE_SIZE = 10

SORT_LABELS = {
    "hp": "💰 Highest Paid",
    "lj": "📅 Latest Join",
    "es": "⏳ Expiring Soon",
    "ex": "❌ Expired",
}

SORT_TITLES = {
    "hp": "Highest Paid",
    "lj": "Latest Join",
    "es": "Expiring Soon",
    "ex": "Expired",
}


# =====================================================
# HELPERS
# =====================================================

def _tg_link(user: User) -> str:
    if user.username:
        return f"https://t.me/{user.username}"
    return f"tg://user?id={user.telegram_id}"


def _main_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="💰 Highest Paid", callback_data="mb_hp_0_0"),
            InlineKeyboardButton(text="📅 Latest Join",  callback_data="mb_lj_0_0"),
        ],
        [
            InlineKeyboardButton(text="⏳ Expiring Soon", callback_data="mb_es_0_0"),
            InlineKeyboardButton(text="❌ Expired",       callback_data="mb_ex_0_0"),
        ],
        [InlineKeyboardButton(text="📺 Filter by Channel", callback_data="mb_ch_menu")],
    ])


async def _fetch_members(sort: str, ch_id: int, page: int):
    """Fetch paginated members based on sort and channel filter."""
    now = datetime.now(timezone.utc)
    offset = page * PAGE_SIZE

    async with async_session() as session:
        base = (
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
        )

        if ch_id != 0:
            base = base.where(Membership.channel_id == ch_id)

        if sort == "hp":
            base = base.where(Membership.is_active == True).order_by(User.highest_amount_paid.desc())
        elif sort == "lj":
            base = base.where(Membership.is_active == True).order_by(Membership.created_at.desc())
        elif sort == "es":
            base = base.where(
                Membership.is_active == True,
                Membership.expiry_date > now
            ).order_by(Membership.expiry_date.asc())
        elif sort == "ex":
            base = base.where(
                Membership.expiry_date < now
            ).order_by(Membership.expiry_date.desc())

        # Total count
        count_result = await session.execute(
            select(func.count()).select_from(base.subquery())
        )
        total = count_result.scalar() or 0

        # Paginated results
        result = await session.execute(base.offset(offset).limit(PAGE_SIZE))
        members = result.all()

    return members, total


async def _fetch_all_members(sort: str, ch_id: int):
    """Fetch ALL members for export (no pagination)."""
    now = datetime.now(timezone.utc)

    async with async_session() as session:
        base = (
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
        )

        if ch_id != 0:
            base = base.where(Membership.channel_id == ch_id)

        if sort == "hp":
            base = base.where(Membership.is_active == True).order_by(User.highest_amount_paid.desc())
        elif sort == "lj":
            base = base.where(Membership.is_active == True).order_by(Membership.created_at.desc())
        elif sort == "es":
            base = base.where(
                Membership.is_active == True,
                Membership.expiry_date > now
            ).order_by(Membership.expiry_date.asc())
        elif sort == "ex":
            base = base.where(
                Membership.expiry_date < now
            ).order_by(Membership.expiry_date.desc())

        result = await session.execute(base)
        return result.all()


async def _show_members(message, sort: str, ch_id: int, page: int, edit: bool = True):
    """Build and send/edit members list message."""
    now = datetime.now(timezone.utc)
    members, total = await _fetch_members(sort, ch_id, page)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)

    ch_label = ""
    if ch_id != 0:
        async with async_session() as session:
            ch = await session.get(Channel, ch_id)
            ch_label = f" — {ch.name}" if ch else ""

    sort_label = SORT_LABELS.get(sort, "Members")
    msg = f"👥 <b>{sort_label}{ch_label}</b>\n"
    msg += f"📊 {total} members | Page {page+1}/{total_pages}\n"
    msg += "━━━━━━━━━━━━━━━━━━━━\n\n"

    message_buttons = []

    if not members:
        msg += "No members found."
    else:
        for idx, (m, u, ch) in enumerate(members, page * PAGE_SIZE + 1):
            name = u.full_name or "No Name"
            username = f"@{u.username}" if u.username else "No username"

            expiry_tz = m.expiry_date
            if expiry_tz.tzinfo is None:
                expiry_tz = expiry_tz.replace(tzinfo=timezone.utc)

            if sort == "ex":
                date_str = f"Expired: {expiry_tz.astimezone(IST).strftime('%d %b %Y')}"
            else:
                days_left = (expiry_tz - now).days
                date_str = f"Expires: {expiry_tz.astimezone(IST).strftime('%d %b %Y')} ({days_left}d)"

            msg += (
                f"{idx}. <b>{name}</b> ({username})\n"
                f"   💰 ₹{int(float(u.highest_amount_paid or 0))} | 📺 {ch.name}\n"
                f"   📅 {date_str}\n\n"
            )

            if u.username:
                message_buttons.append(InlineKeyboardButton(
                    text=f"💬 {name[:15]}",
                    url=f"https://t.me/{u.username}"
                ))
            message_buttons.append(InlineKeyboardButton(
                text="👤 Info",
                callback_data=f"mb_info_{u.telegram_id}"
            ))

    # Build keyboard — 💬 + 👤 per row (one user per row)
    keyboard_rows = []
    for i in range(0, len(message_buttons), 2):
        keyboard_rows.append(message_buttons[i:i+2])

    # Pagination row
    nav_row = []
    if page > 0:
        nav_row.append(InlineKeyboardButton(text="⬅️ Prev", callback_data=f"mb_{sort}_{ch_id}_{page-1}"))
    nav_row.append(InlineKeyboardButton(text=f"{page+1}/{total_pages}", callback_data="mb_noop"))
    if page + 1 < total_pages:
        nav_row.append(InlineKeyboardButton(text="Next ➡️", callback_data=f"mb_{sort}_{ch_id}_{page+1}"))
    keyboard_rows.append(nav_row)

    # Export + Back row
    keyboard_rows.append([
        InlineKeyboardButton(text="📥 Export View", callback_data=f"mb_exp_{sort}_{ch_id}"),
        InlineKeyboardButton(text="🔙 Back", callback_data="mb_menu"),
    ])

    reply_markup = InlineKeyboardMarkup(inline_keyboard=keyboard_rows)

    try:
        if edit:
            await message.edit_text(msg, parse_mode="HTML", reply_markup=reply_markup)
        else:
            await message.answer(msg, parse_mode="HTML", reply_markup=reply_markup)
    except Exception:
        await message.answer(msg[:4000], parse_mode="HTML", reply_markup=reply_markup)


# =====================================================
# /MEMBERS COMMAND
# =====================================================

@router.message(Command("members"))
async def members_command(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("⛔ Admin only.")
        return
    await message.answer(
        "👥 <b>Members Panel</b>\n\nSelect how to view members:",
        reply_markup=_main_keyboard(),
        parse_mode="HTML"
    )


# =====================================================
# BACK TO MEMBERS MENU
# =====================================================

@router.callback_query(F.data == "mb_menu")
async def back_to_members_menu(callback: CallbackQuery):
    try:
        await callback.answer()
    except Exception:
        pass
    await callback.message.edit_text(
        "👥 <b>Members Panel</b>\n\nSelect how to view members:",
        reply_markup=_main_keyboard(),
        parse_mode="HTML"
    )


# =====================================================
# CHANNEL FILTER MENU
# =====================================================

@router.callback_query(F.data == "mb_ch_menu")
async def channel_filter_menu(callback: CallbackQuery):
    try:
        await callback.answer()
    except Exception:
        pass

    async with async_session() as session:
        result = await session.execute(
            select(Channel).where(Channel.is_active == True).order_by(Channel.id)
        )
        channels = result.scalars().all()

    keyboard = []
    for ch in channels:
        keyboard.append([InlineKeyboardButton(
            text=ch.name,
            callback_data=f"mb_chsel_{ch.id}"
        )])
    keyboard.append([InlineKeyboardButton(text="🔙 Back", callback_data="mb_menu")])

    await callback.message.edit_text(
        "📺 <b>Filter by Channel</b>\n\nSelect a channel:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=keyboard),
        parse_mode="HTML"
    )


@router.callback_query(F.data.startswith("mb_chsel_"))
async def channel_filter_selected(callback: CallbackQuery):
    try:
        await callback.answer()
    except Exception:
        pass

    ch_id = int(callback.data.split("_")[2])

    async with async_session() as session:
        ch = await session.get(Channel, ch_id)

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="💰 Highest Paid", callback_data=f"mb_hp_{ch_id}_0"),
            InlineKeyboardButton(text="📅 Latest Join",  callback_data=f"mb_lj_{ch_id}_0"),
        ],
        [
            InlineKeyboardButton(text="⏳ Expiring Soon", callback_data=f"mb_es_{ch_id}_0"),
            InlineKeyboardButton(text="❌ Expired",       callback_data=f"mb_ex_{ch_id}_0"),
        ],
        [InlineKeyboardButton(text="🔙 Back", callback_data="mb_ch_menu")],
    ])

    await callback.message.edit_text(
        f"📺 <b>{ch.name}</b>\n\nSelect sort order:",
        reply_markup=keyboard,
        parse_mode="HTML"
    )


# =====================================================
# SORT / PAGINATION CALLBACKS
# =====================================================

@router.callback_query(F.data.regexp(r"^mb_(hp|lj|es|ex)_\d+_\d+$"))
async def handle_members_view(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("⛔ Admin only.", show_alert=True)
        return
    try:
        await callback.answer()
    except Exception:
        pass

    parts = callback.data.split("_")
    sort = parts[1]
    ch_id = int(parts[2])
    page = int(parts[3])

    await _show_members(callback.message, sort, ch_id, page, edit=True)


@router.callback_query(F.data == "mb_noop")
async def noop(callback: CallbackQuery):
    try:
        await callback.answer()
    except Exception:
        pass


# =====================================================
# USER INFO BUTTON
# =====================================================

@router.callback_query(F.data.startswith("mb_info_"))
async def member_info(callback: CallbackQuery):
    try:
        await callback.answer()
    except Exception:
        pass
    telegram_id = int(callback.data.split("_")[2])
    from backend.app.bot.handlers.admin_panel import _do_user_info
    await _do_user_info(callback.message, telegram_id)


# =====================================================
# EXPORT CURRENT VIEW
# =====================================================

@router.callback_query(F.data.startswith("mb_exp_"))
async def export_view(callback: CallbackQuery):
    try:
        await callback.answer("⏳ Generating Excel...")
    except Exception:
        pass

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    parts = callback.data.split("_")
    sort = parts[2]
    ch_id = int(parts[3])

    now = datetime.now(timezone.utc)
    all_members = await _fetch_all_members(sort, ch_id)

    wb = Workbook()
    ws = wb.active
    ws.title = SORT_TITLES.get(sort, "Members")

    headers = [
        "Name", "Username", "Telegram ID", "Channel",
        "Amount Paid (₹)", "Highest Paid (₹)",
        "Expiry Date", "Days Left", "Telegram Link"
    ]

    # Style header
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20

    for m, u, ch in all_members:
        expiry_tz = m.expiry_date
        if expiry_tz.tzinfo is None:
            expiry_tz = expiry_tz.replace(tzinfo=timezone.utc)
        days_left = (expiry_tz - now).days
        ws.append([
            u.full_name or "N/A",
            f"@{u.username}" if u.username else "N/A",
            u.telegram_id,
            ch.name,
            float(m.amount_paid),
            float(u.highest_amount_paid or 0),
            expiry_tz.astimezone(IST).strftime("%d %b %Y"),
            days_left,
            _tg_link(u)
        ])

    # Auto-size columns
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"members_{sort}_{datetime.now(IST).strftime('%Y-%m-%d')}.xlsx"

    await callback.message.answer_document(
        document=BufferedInputFile(buffer.read(), filename=filename),
        caption=(
            f"📥 <b>Export: {SORT_TITLES.get(sort, 'Members')}</b>\n"
            f"📊 {len(all_members)} members"
        ),
        parse_mode="HTML"
    )