import io
import os
from datetime import datetime, timedelta, timezone
from sqlalchemy import select, func

from backend.app.db.session import async_session
from backend.app.db.models import User, Payment, Membership, Channel
from backend.bot.bot import bot
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, BufferedInputFile


# =========================
# CONFIG
# =========================

ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x]
IST = timezone(timedelta(hours=5, minutes=30))


# =========================
# HELPERS
# =========================

async def _send_to_admins(text: str):
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text, parse_mode="HTML")
        except Exception as e:
            print(f"❌ Failed to send report to {admin_id}: {e}")


async def _all_channels_revenue(session, start, end):
    rows = await session.execute(
        select(
            Channel.name,
            func.coalesce(func.sum(Payment.amount), 0)
        )
        .outerjoin(
            Payment,
            (Payment.channel_id == Channel.id)
            & (Payment.status == "captured")
            & (Payment.created_at >= start)
            & (Payment.created_at < end)
        )
        .group_by(Channel.name)
        .order_by(func.coalesce(func.sum(Payment.amount), 0).desc())
    )
    return rows.all()


def _tg_link(user: User) -> str:
    if user.username:
        return f"https://t.me/{user.username}"
    return ""


def _style_header(ws, headers: list):
    from openpyxl.styles import Font, PatternFill, Alignment
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 20


def _autosize_columns(ws):
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)


# =========================
# DAILY REPORT (YESTERDAY)
# =========================

async def send_daily_report():
    now = datetime.now(timezone.utc)

    report_day = now.date() - timedelta(days=1)
    start = datetime.combine(report_day, datetime.min.time(), tzinfo=timezone.utc)
    end = start + timedelta(days=1)

    prev_start = start - timedelta(days=1)
    prev_end = start

    async with async_session() as session:

        revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(start, end))
        )

        prev_revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(prev_start, prev_end))
        )

        change = ((revenue - prev_revenue) / prev_revenue * 100) if prev_revenue else 0

        new_users = await session.scalar(
            select(func.count(User.id))
            .where(User.created_at.between(start, end))
        )

        new_subs = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.created_at.between(start, end))
        )

        expired = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.expiry_date.between(start, end))
        )

        failed = await session.scalar(
            select(func.count(Payment.id))
            .where(Payment.status == "failed",
                   Payment.created_at.between(start, end))
        )

        active = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.is_active == True,
                   Membership.expiry_date > now)
        )

        channels = await _all_channels_revenue(session, start, end)

    date_label = start.astimezone(IST).strftime("%d %b")

    channel_text = "\n".join(
        f"• {name} – ₹{int(amount)}" for name, amount in channels
    ) or "• No channels"

    msg = (
        f"📊 <b>Daily Report ({date_label})</b>\n\n"
        f"💰 Revenue: ₹{int(revenue)}\n"
        f"📈 Change vs prev day: {change:+.1f}%\n\n"
        f"🆕 New users: {new_users}\n"
        f"🆕 New subs: {new_subs}\n\n"
        f"❌ Expired subs: {expired}\n"
        f"❌ Failed payments: {failed}\n\n"
        f"📺 <b>Channel-wise Revenue:</b>\n{channel_text}\n\n"
        f"✅ Active subs: {active}"
    )

    await _send_to_admins(msg)


# =========================
# WEEKLY REPORT (LAST 7 DAYS)
# =========================

async def send_weekly_report():
    now = datetime.now(timezone.utc)

    end = datetime.combine(now.date(), datetime.min.time(), tzinfo=timezone.utc)
    start = end - timedelta(days=7)

    prev_start = start - timedelta(days=7)
    prev_end = start

    async with async_session() as session:

        revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(start, end))
        )

        prev_revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(prev_start, prev_end))
        )

        change = ((revenue - prev_revenue) / prev_revenue * 100) if prev_revenue else 0

        users = await session.scalar(
            select(func.count(User.id))
            .where(User.created_at.between(start, end))
        )

        subs = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.created_at.between(start, end))
        )

        expired = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.expiry_date.between(start, end))
        )

        failed = await session.scalar(
            select(func.count(Payment.id))
            .where(Payment.status == "failed",
                   Payment.created_at.between(start, end))
        )

        active = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.is_active == True,
                   Membership.expiry_date > now)
        )

        channels = await _all_channels_revenue(session, start, end)

    range_label = f"{start.astimezone(IST):%d %b} – {(end - timedelta(days=1)).astimezone(IST):%d %b}"

    channel_text = "\n".join(
        f"• {name} – ₹{int(amount)}" for name, amount in channels
    ) or "• No channels"

    msg = (
        f"📊 <b>Weekly Summary ({range_label})</b>\n\n"
        f"💰 Revenue: ₹{int(revenue)}\n"
        f"📈 Change vs prev week: {change:+.1f}%\n\n"
        f"🆕 New users: {users}\n"
        f"🆕 New subs: {subs}\n\n"
        f"❌ Expired subs: {expired}\n"
        f"❌ Failed payments: {failed}\n\n"
        f"📺 <b>Channel-wise Revenue:</b>\n{channel_text}\n\n"
        f"✅ Active subs: {active}"
    )

    await _send_to_admins(msg)


# =========================
# MONTHLY REPORT (PREVIOUS MONTH)
# =========================

async def send_monthly_report():
    now = datetime.now(timezone.utc)

    first_this_month = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    end = first_this_month
    start = (first_this_month - timedelta(days=1)).replace(day=1)

    prev_start = (start - timedelta(days=1)).replace(day=1)
    prev_end = start

    async with async_session() as session:

        revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(start, end))
        )

        prev_revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(prev_start, prev_end))
        )

        change = ((revenue - prev_revenue) / prev_revenue * 100) if prev_revenue else 0

        users = await session.scalar(
            select(func.count(User.id))
            .where(User.created_at.between(start, end))
        )

        subs = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.created_at.between(start, end))
        )

        expired = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.expiry_date.between(start, end))
        )

        failed = await session.scalar(
            select(func.count(Payment.id))
            .where(Payment.status == "failed",
                   Payment.created_at.between(start, end))
        )

        active = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.is_active == True,
                   Membership.expiry_date > end)
        )

        channels = await _all_channels_revenue(session, start, end)

    month_label = start.astimezone(IST).strftime("%b %Y")

    channel_text = "\n".join(
        f"• {name} – ₹{int(amount)}" for name, amount in channels
    ) or "• No channels"

    msg = (
        f"📊 <b>Monthly Summary ({month_label})</b>\n\n"
        f"💰 Revenue: ₹{int(revenue)}\n"
        f"📈 Change vs prev month: {change:+.1f}%\n\n"
        f"🆕 New users: {users}\n"
        f"🆕 New subs: {subs}\n\n"
        f"❌ Expired subs: {expired}\n"
        f"❌ Failed payments: {failed}\n\n"
        f"📺 <b>Channel-wise Revenue:</b>\n{channel_text}\n\n"
        f"✅ Active subs (month end): {active}"
    )

    await _send_to_admins(msg)


# =========================
# YEARLY REPORT (PREVIOUS YEAR)
# =========================

async def send_yearly_report():
    now = datetime.now(timezone.utc)

    start = datetime(now.year - 1, 1, 1, tzinfo=timezone.utc)
    end = datetime(now.year, 1, 1, tzinfo=timezone.utc)

    prev_start = datetime(now.year - 2, 1, 1, tzinfo=timezone.utc)
    prev_end = start

    async with async_session() as session:

        revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(start, end))
        )

        prev_revenue = await session.scalar(
            select(func.coalesce(func.sum(Payment.amount), 0))
            .where(Payment.status == "captured",
                   Payment.created_at.between(prev_start, prev_end))
        )

        change = ((revenue - prev_revenue) / prev_revenue * 100) if prev_revenue else 0

        users = await session.scalar(
            select(func.count(User.id))
            .where(User.created_at.between(start, end))
        )

        subs = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.created_at.between(start, end))
        )

        expired = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.expiry_date.between(start, end))
        )

        failed = await session.scalar(
            select(func.count(Payment.id))
            .where(Payment.status == "failed",
                   Payment.created_at.between(start, end))
        )

        active = await session.scalar(
            select(func.count(Membership.id))
            .where(Membership.is_active == True,
                   Membership.expiry_date > end)
        )

        channels = await _all_channels_revenue(session, start, end)

    channel_text = "\n".join(
        f"• {name} – ₹{int(amount)}" for name, amount in channels
    ) or "• No channels"

    msg = (
        f"📊 <b>Yearly Summary ({start.year})</b>\n\n"
        f"💰 Revenue: ₹{int(revenue)}\n"
        f"📈 Change vs prev year: {change:+.1f}%\n\n"
        f"🆕 New users: {users}\n"
        f"🆕 New subs: {subs}\n\n"
        f"❌ Expired subs: {expired}\n"
        f"❌ Failed payments: {failed}\n\n"
        f"📺 <b>Channel-wise Revenue:</b>\n{channel_text}\n\n"
        f"✅ Active subs (year end): {active}"
    )

    await _send_to_admins(msg)


# =========================
# MEMBER DAILY REPORT
# =========================

async def send_member_daily_report(date=None):
    now = datetime.now(timezone.utc)

    if date is None:
        report_day = now.date() - timedelta(days=1)
    else:
        report_day = date

    start = datetime.combine(report_day, datetime.min.time(), tzinfo=timezone.utc)
    end = start + timedelta(days=1)

    today_start = datetime.combine(now.date(), datetime.min.time(), tzinfo=timezone.utc)
    today_end = today_start + timedelta(days=1)

    async with async_session() as session:

        new_result = await session.execute(
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
            .where(Membership.created_at.between(start, end))
            .order_by(Membership.created_at.desc())
        )
        new_members = new_result.all()

        expiring_result = await session.execute(
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
            .where(
                Membership.is_active == True,
                Membership.expiry_date.between(today_start, today_end)
            )
            .order_by(Membership.expiry_date.asc())
        )
        expiring = expiring_result.all()

        expired_result = await session.execute(
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
            .where(Membership.expiry_date.between(start, end))
            .order_by(Membership.expiry_date.desc())
        )
        expired = expired_result.all()

    date_label = start.astimezone(IST).strftime("%d %b %Y")

    msg = (
        f"👥 <b>Member Report — {date_label}</b>\n"
        f"📊 {len(new_members)} new | {len(expiring)} expiring today | {len(expired)} expired\n"
        f"━━━━━━━━━━━━━━━━━━━━\n\n"
    )

    keyboards = []

    if new_members:
        msg += "🆕 <b>NEW MEMBERS</b>\n\n"
        for idx, (m, u, ch) in enumerate(new_members, 1):
            name = u.full_name or "No Name"
            username = f"@{u.username}" if u.username else "No username"
            msg += (
                f"{idx}. <b>{name}</b> ({username})\n"
                f"   📺 {ch.name} | ₹{int(float(m.amount_paid))} | {m.validity_days}d\n\n"
            )
            if u.username:
                keyboards.append(InlineKeyboardButton(
                    text=f"💬 {name[:15]}",
                    url=f"https://t.me/{u.username}"
                ))
        msg += "━━━━━━━━━━━━━━━━━━━━\n\n"
    else:
        msg += "🆕 <b>NEW MEMBERS</b>\nNone\n━━━━━━━━━━━━━━━━━━━━\n\n"

    if expiring:
        msg += "⚠️ <b>EXPIRING TODAY</b>\n\n"
        for idx, (m, u, ch) in enumerate(expiring, 1):
            name = u.full_name or "No Name"
            username = f"@{u.username}" if u.username else "No username"
            expiry_time = m.expiry_date.astimezone(IST).strftime("%I:%M %p")
            msg += (
                f"{idx}. <b>{name}</b> ({username})\n"
                f"   📺 {ch.name} | Expires at {expiry_time} IST\n\n"
            )
            if u.username:
                keyboards.append(InlineKeyboardButton(
                    text=f"💬 {name[:15]}",
                    url=f"https://t.me/{u.username}"
                ))
        msg += "━━━━━━━━━━━━━━━━━━━━\n\n"
    else:
        msg += "⚠️ <b>EXPIRING TODAY</b>\nNone\n━━━━━━━━━━━━━━━━━━━━\n\n"

    if expired:
        msg += "❌ <b>EXPIRED</b>\n\n"
        for idx, (m, u, ch) in enumerate(expired, 1):
            name = u.full_name or "No Name"
            username = f"@{u.username}" if u.username else "No username"
            expiry_time = m.expiry_date.astimezone(IST).strftime("%I:%M %p")
            msg += (
                f"{idx}. <b>{name}</b> ({username})\n"
                f"   📺 {ch.name} | Expired at {expiry_time} IST\n\n"
            )
            if u.username:
                keyboards.append(InlineKeyboardButton(
                    text=f"💬 {name[:15]}",
                    url=f"https://t.me/{u.username}"
                ))
    else:
        msg += "❌ <b>EXPIRED</b>\nNone\n"

    keyboard_rows = [keyboards[i:i+2] for i in range(0, len(keyboards), 2)]
    reply_markup = InlineKeyboardMarkup(inline_keyboard=keyboard_rows) if keyboard_rows else None

    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(
                admin_id, msg,
                parse_mode="HTML",
                reply_markup=reply_markup
            )
        except Exception as e:
            print(f"❌ Failed to send member report to {admin_id}: {e}")


# =========================
# EXCEL DAILY REPORT
# =========================

async def send_excel_report():
    from openpyxl import Workbook

    now = datetime.now(timezone.utc)
    report_day = now.date() - timedelta(days=1)
    start = datetime.combine(report_day, datetime.min.time(), tzinfo=timezone.utc)
    end = start + timedelta(days=1)
    today_start = datetime.combine(now.date(), datetime.min.time(), tzinfo=timezone.utc)
    today_end = today_start + timedelta(days=1)

    async with async_session() as session:

        # Sheet 1 — New Members Yesterday
        new_result = await session.execute(
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
            .where(Membership.created_at.between(start, end))
            .order_by(Membership.created_at.desc())
        )
        new_members = new_result.all()

        # Sheet 2 — Expiring Today
        expiring_result = await session.execute(
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
            .where(
                Membership.is_active == True,
                Membership.expiry_date.between(today_start, today_end)
            )
            .order_by(Membership.expiry_date.asc())
        )
        expiring = expiring_result.all()

        # Sheet 3 — Expired Yesterday
        expired_result = await session.execute(
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
            .where(Membership.expiry_date.between(start, end))
            .order_by(Membership.expiry_date.desc())
        )
        expired = expired_result.all()

        # Sheet 4 — All Active Members
        active_result = await session.execute(
            select(Membership, User, Channel)
            .join(User, Membership.user_id == User.id)
            .join(Channel, Membership.channel_id == Channel.id)
            .where(
                Membership.is_active == True,
                Membership.expiry_date > now
            )
            .order_by(User.highest_amount_paid.desc())
        )
        active_members = active_result.all()

    wb = Workbook()

    # ── Sheet 1: New Members ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "New Members"
    headers1 = ["Name", "Username", "Telegram ID", "Channel", "Plan (Days)", "Amount (₹)", "Join Date", "Telegram Link"]
    _style_header(ws1, headers1)
    for m, u, ch in new_members:
        ws1.append([
            u.full_name or "N/A",
            f"@{u.username}" if u.username else "N/A",
            u.telegram_id,
            ch.name,
            m.validity_days,
            float(m.amount_paid),
            m.created_at.astimezone(IST).strftime("%d %b %Y %I:%M %p") if m.created_at else "N/A",
            _tg_link(u)
        ])
    _autosize_columns(ws1)

    # ── Sheet 2: Expiring Today ──────────────────────────────────
    ws2 = wb.create_sheet("Expiring Today")
    headers2 = ["Name", "Username", "Telegram ID", "Channel", "Expiry Date", "Expiry Time (IST)", "Amount (₹)", "Telegram Link"]
    _style_header(ws2, headers2)
    for m, u, ch in expiring:
        ws2.append([
            u.full_name or "N/A",
            f"@{u.username}" if u.username else "N/A",
            u.telegram_id,
            ch.name,
            m.expiry_date.astimezone(IST).strftime("%d %b %Y"),
            m.expiry_date.astimezone(IST).strftime("%I:%M %p"),
            float(m.amount_paid),
            _tg_link(u)
        ])
    _autosize_columns(ws2)

    # ── Sheet 3: Expired Yesterday ───────────────────────────────
    ws3 = wb.create_sheet("Expired Yesterday")
    headers3 = ["Name", "Username", "Telegram ID", "Channel", "Expired Date", "Amount (₹)", "Telegram Link"]
    _style_header(ws3, headers3)
    for m, u, ch in expired:
        ws3.append([
            u.full_name or "N/A",
            f"@{u.username}" if u.username else "N/A",
            u.telegram_id,
            ch.name,
            m.expiry_date.astimezone(IST).strftime("%d %b %Y"),
            float(m.amount_paid),
            _tg_link(u)
        ])
    _autosize_columns(ws3)

    # ── Sheet 4: All Active Members ──────────────────────────────
    ws4 = wb.create_sheet("All Active Members")
    headers4 = ["Name", "Username", "Telegram ID", "Channel", "Plan (Days)", "Amount (₹)", "Highest Paid (₹)", "Expiry Date", "Days Left", "Telegram Link"]
    _style_header(ws4, headers4)
    for m, u, ch in active_members:
        expiry_tz = m.expiry_date
        if expiry_tz.tzinfo is None:
            expiry_tz = expiry_tz.replace(tzinfo=timezone.utc)
        days_left = (expiry_tz - now).days
        ws4.append([
            u.full_name or "N/A",
            f"@{u.username}" if u.username else "N/A",
            u.telegram_id,
            ch.name,
            m.validity_days,
            float(m.amount_paid),
            float(u.highest_amount_paid or 0),
            expiry_tz.astimezone(IST).strftime("%d %b %Y"),
            days_left,
            _tg_link(u)
        ])
    _autosize_columns(ws4)

    # ── Save and send ────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"report_{report_day.strftime('%Y-%m-%d')}.xlsx"
    caption = (
        f"📊 <b>Daily Excel Report — {report_day.strftime('%d %b %Y')}</b>\n\n"
        f"📋 {len(new_members)} new members\n"
        f"⚠️ {len(expiring)} expiring today\n"
        f"❌ {len(expired)} expired yesterday\n"
        f"✅ {len(active_members)} total active"
    )

    for admin_id in ADMIN_IDS:
        try:
            buffer.seek(0)
            await bot.send_document(
                chat_id=admin_id,
                document=BufferedInputFile(buffer.read(), filename=filename),
                caption=caption,
                parse_mode="HTML"
            )
        except Exception as e:
            print(f"❌ Failed to send Excel report to {admin_id}: {e}")